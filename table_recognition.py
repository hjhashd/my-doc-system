
from __future__ import annotations
from pydantic_settings import BaseSettings, SettingsConfigDict
import os, re, json, bisect, shutil
from typing import List, Dict, Optional, Tuple
from datetime import datetime
import httpx
import pandas as pd  # NEW
from openpyxl import load_workbook  # NEW
import hashlib  # 新增


# 在工具函数区域添加
def generate_table_id(file_path: str, title: str) -> str:
    """
    根据文件路径和标题生成唯一的表ID
    使用SHA256哈希的前16位
    """
    # 组合 file_path 和 title
    combined = f"{file_path}::{title}"
    # 生成哈希
    hash_obj = hashlib.sha256(combined.encode('utf-8'))
    # 取前16位（32个字符）作为ID
    table_id = hash_obj.hexdigest()[:16]
    return table_id


# =========================
# 配置
# =========================
class Settings(BaseSettings):
    model_config = SettingsConfigDict(env_file=".env", env_file_encoding="utf-8", extra="allow")

    MODEL_PROVIDER: str = "deepseek"
    MODEL_ID: str = "deepseek-reasoner"
    DEEPSEEK_API_KEY: str | None = "sk-d0be6f95235f4a1c981275488c6f64ba"  # 建议用环境变量
    DEEPSEEK_BASE_URL: str = "https://api.deepseek.com"

    OPENAI_API_KEY: str | None = None
    OLLAMA_BASE_URL: str | None = "http://localhost:11434"
    GEMINI_API_KEY: str | None = None

    TIMEOUT_S: int = 6000
    OUTPUT_DIR: str = "./out"

    # ---------- 新增：附件目录/预览预算 ----------
    ATTACHMENT_OUT_DIR: str = "/data/cwd_cq/out"  # NEW
    ATTACHMENT_SEARCH_DIRS: List[str] = ["/data/cwd_cq/out", "."]  # NEW
    SHEET_PREVIEW_ROWS: int = 40  # NEW
    SHEET_PREVIEW_COLS: int = 32  # NEW

    # ========== 新增：路径参数 ==========
    INPUT_FILE_PATH: str = ""  # 基础输入路径
    AGENT_USER_ID: str = ""    # 用户ID
    TASK_ID: str = ""          # 任务ID
    # ==================================

    LLM_MAX_PREDICT: int = 2048
    TITLE_CONTEXT_LINES: int = 16
    TITLE_CONTEXT_CHARS: int = 1200


settings = Settings()
os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
os.makedirs(settings.ATTACHMENT_OUT_DIR, exist_ok=True)  # NEW

# =========================
# 工具函数（保留 + 新增）
# =========================
_CH_SEP_LINE = set("|-: +=")


def _is_visual_separator(s: str) -> bool:
    s = (s or "").strip()
    if not s:
        return True
    if '|' in s:
        return False
    return set(s) <= _CH_SEP_LINE


def _add_line_numbers(lines: List[str], begin_0based: int) -> List[str]:
    return [f"L{begin_0based + i + 1}: {ln}" for i, ln in enumerate(lines)]


def _now_suffix() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ---------- NEW: 在全文中找 *.xlsx ----------
_XLSX_RE = re.compile(r'([A-Za-z0-9_\-（）()\u4e00-\u9fa5\.]+\.xlsx)\b')


def find_xlsx_mentions(text: str) -> List[Dict]:
    lines = text.splitlines()
    hits = []
    for idx, ln in enumerate(lines):
        for m in _XLSX_RE.finditer(ln):
            hits.append({"file_name": m.group(1), "line_idx": idx})
    return hits


# ---------- NEW: 附件定位/落盘 ----------
def resolve_and_stage_attachment(name: str, input_base_path: str = "",
                                     agent_user_id: str = "", task_id: str = "") -> Optional[str]:
    """
    附件定位/落盘，支持路径拼接

    Args:
        name: 文件名或相对路径
        input_base_path: 基础输入路径
        agent_user_id: 用户ID
        task_id: 任务ID
    """
    # 如果提供了路径参数，先尝试拼接路径
    if input_base_path and agent_user_id and task_id:
        # 去掉路径开头的 / 或 \ (避免os.path.join把它当作绝对路径)
        clean_name = name.lstrip('/\\')

        # 拼接完整路径: input_base_path/agentUserId/task_id/相对路径
        full_path = os.path.join(
            input_base_path,
            str(agent_user_id),
            str(task_id),
            clean_name
        )
        full_path = os.path.normpath(full_path)

        # 如果拼接后的路径存在
        if os.path.exists(full_path):
            dst = os.path.join(settings.ATTACHMENT_OUT_DIR, os.path.basename(name))
            if os.path.abspath(full_path) != os.path.abspath(dst):
                shutil.copy2(full_path, dst)
            print(f"  ✓ 找到文件（拼接路径）: {full_path}")
            return dst

    # 原有逻辑：已是绝对路径且存在
    if os.path.isabs(name) and os.path.exists(name):
        dst = os.path.join(settings.ATTACHMENT_OUT_DIR, os.path.basename(name))
        if os.path.abspath(name) != os.path.abspath(dst):
            shutil.copy2(name, dst)
        print(f"  ✓ 找到文件（绝对路径）: {name}")
        return dst

    # 原有逻辑：在搜索目录中查找
    for d in settings.ATTACHMENT_SEARCH_DIRS:
        cand = os.path.join(d, name)
        if os.path.exists(cand):
            dst = os.path.join(settings.ATTACHMENT_OUT_DIR, os.path.basename(name))
            if os.path.abspath(cand) != os.path.abspath(dst):
                shutil.copy2(cand, dst)
            print(f"  ✓ 找到文件（搜索目录）: {cand}")
            return dst

    print(f"  ✗ 文件未找到: {name}")
    return None


# ---------- NEW: Excel 读取并展开合并 ----------
def read_xlsx_expand_merged(path: str, sheet) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[sheet]] if isinstance(sheet, int) else wb[sheet]
    max_row, max_col = ws.max_row, ws.max_column
    grid = [[ws.cell(r, c).value for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)]
    for rng in list(ws.merged_cells.ranges):
        r1, c1, r2, c2 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        v = grid[r1 - 1][c1 - 1]
        for r in range(r1 - 1, r2):
            for c in range(c1 - 1, c2):
                if grid[r][c] in (None, ""):
                    grid[r][c] = v
    return pd.DataFrame(grid)


# ---------- NEW: DataFrame 转 markdown 预览行 ----------
def df_to_md_lines(df: pd.DataFrame, max_rows=40, max_cols=32) -> List[str]:
    dfv = df.iloc[:max_rows, :max_cols].fillna("")
    lines = []
    for _, row in dfv.iterrows():
        cells = [str(x).strip() for x in row.tolist()]
        lines.append("| " + " | ".join(cells) + " |")
    return lines


# ---------- NEW: 切列（子表） ----------
def slice_df_columns(df: pd.DataFrame, col_range: List[int]) -> pd.DataFrame:
    s, e = col_range
    s0, e0 = max(1, s) - 1, min(df.shape[1], e)
    return df.iloc[:, s0:e0]


# =========================
# LLM 客户端（与你现有的一致）
# =========================
class LLMClient:
    def __init__(self, s: Settings):
        self.s = s

    def _clean_json_response(self, response: str) -> str:
        import re
        if '<think>' in response and '</think>' in response:
            response = re.sub(r'<think>.*?</think>', '', response, flags=re.DOTALL)
        response = (response or "").strip()
        m = re.search(r'```json\s*(.*?)\s*```', response, re.DOTALL)
        if m: return m.group(1)
        m = re.search(r'```\s*(.*?)\s*```', response, re.DOTALL)
        return m.group(1) if m else response

    def _call_deepseek(self, prompt: str) -> str:
        base = self.s.DEEPSEEK_BASE_URL.rstrip("/")
        url = f"{base}/chat/completions"
        headers = {"Authorization": f"Bearer {self.s.DEEPSEEK_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": self.s.MODEL_ID,
            "messages": [
                {"role": "system", "content":
                    "You are an expert at document tables: detect titles, units, dates, composite layout, and multi-level headers. Output JSON only."},
                {"role": "user", "content": prompt},
            ],
            "stream": False,
            "response_format": {"type": "json_object"},
        }
        if self.s.MODEL_ID == "deepseek-chat":
            payload.update({"temperature": 0.1, "top_p": 0.95})
        with httpx.Client(timeout=self.s.TIMEOUT_S) as client:
            resp = client.post(url, headers=headers, json=payload)
            resp.raise_for_status()
            data = resp.json()
            msg = data["choices"][0]["message"]
            content = (msg.get("content") or "").strip()
            if not content and msg.get("reasoning_content"):
                import re
                m = re.search(r"\{[\s\S]*\}", msg["reasoning_content"])
                if m: content = m.group(0)
            return content

    def call(self, prompt: str) -> Dict:
        raw = self._call_deepseek(prompt) if self.s.MODEL_PROVIDER == "deepseek" else ""
        cleaned = self._clean_json_response(raw)
        try:
            return json.loads(cleaned)
        except Exception:
            return {}


# =========================
# 主解析器（新增 xlsx 分支）
# =========================
class MultiHeaderTableExtractor:
    def __init__(self):
        self.s = settings
        self.llm = LLMClient(self.s)

    # def _get_context_before(self, lines: List[str], table_start_idx: int) -> Dict:
    #     """
    #     获取表格之前的上下文，遇到另一个.xlsx文件时停止
    #     """
    #     end = max(table_start_idx - 1, 0)
    #     max_lines = self.s.TITLE_CONTEXT_LINES
    #     max_chars = self.s.TITLE_CONTEXT_CHARS
    #     theoretical_begin = max(0, end - max_lines + 1)
    #     actual_begin = end + 1
    #
    #     # 从当前位置向上查找，直到遇到另一个xlsx文件或达到限制
    #     for i in range(end, theoretical_begin - 1, -1):
    #         if i < 0:
    #             actual_begin = 0
    #             break
    #
    #         # 检查当前行是否包含xlsx文件（使用同样的正则表达式）
    #         if i < len(lines) and i != table_start_idx:  # 排除当前表格所在行
    #             if _XLSX_RE.search(lines[i]):  # 如果找到另一个xlsx文件
    #                 actual_begin = i + 1  # 从这个xlsx文件的下一行开始
    #                 break
    #
    #         actual_begin = i
    #
    #     actual_begin = max(0, min(actual_begin, end + 1))
    #     window_lines = lines[actual_begin:end + 1]
    #
    #     # 处理字符长度限制
    #     joined = "\n".join(window_lines)
    #     if len(joined) > max_chars:
    #         # 如果超过字符限制，从后往前截取
    #         joined = joined[-max_chars:]
    #         tmp = joined.split("\n")
    #         if len(tmp) > 1 and not joined.startswith(tmp[0]):
    #             tmp = tmp[1:]  # 去掉不完整的第一行
    #         window_lines = tmp
    #         actual_begin = end - len(window_lines) + 1
    #
    #     return {
    #         "lines": window_lines,
    #         "begin": actual_begin,
    #         "end": end,
    #         "numbered": _add_line_numbers(window_lines, actual_begin)
    #     }

    def _get_context_before(self, lines: List[str], table_start_idx: int) -> Dict:
        """
        获取表格之前的上下文，智能跳过连续的xlsx文件引用
        优化：当遇到连续的xlsx文件时，继续向上查找共同的标题
        """
        end = max(table_start_idx - 1, 0)
        max_lines = self.s.TITLE_CONTEXT_LINES
        max_chars = self.s.TITLE_CONTEXT_CHARS
        theoretical_begin = max(0, end - max_lines + 1)

        # 第一阶段：向上跳过连续的xlsx文件和空行
        # 目的：找到最近的"有实质内容"的行
        i = end
        while i >= theoretical_begin:
            if i < 0:
                break
            if i >= len(lines):
                i -= 1
                continue

            # 跳过当前表格所在行
            if i == table_start_idx:
                i -= 1
                continue

            line = lines[i].strip()

            # 如果是xlsx文件行或空行，继续向上跳过
            if _XLSX_RE.search(lines[i]) or not line:
                i -= 1
                continue

            # 遇到有内容的行，停止跳过
            break

        # i 现在指向最近的有内容的非xlsx行
        context_end = max(i, 0)

        # 第二阶段：从有内容的行开始，正常向上查找上下文
        # 遇到更早的xlsx文件时才停止（说明是其他表格的区域）
        actual_begin = context_end + 1

        for j in range(context_end, max(0, context_end - max_lines + 1) - 1, -1):
            if j < 0:
                actual_begin = 0
                break

            if j >= len(lines):
                actual_begin = max(0, j)
                continue

            # 如果遇到xlsx文件（且不在刚才跳过的连续区域内），说明到了其他表格区域，停止
            if _XLSX_RE.search(lines[j]) and j < end - 1:
                actual_begin = j + 1
                break

            actual_begin = j

        # 确保范围有效
        actual_begin = max(0, min(actual_begin, context_end + 1))
        window_lines = lines[actual_begin:context_end + 1]

        # 处理字符长度限制
        joined = "\n".join(window_lines)
        if len(joined) > max_chars:
            joined = joined[-max_chars:]
            tmp = joined.split("\n")
            if len(tmp) > 1 and not joined.startswith(tmp[0]):
                tmp = tmp[1:]
            window_lines = tmp
            actual_begin = context_end - len(window_lines) + 1

        return {
            "lines": window_lines,
            "begin": actual_begin,
            "end": context_end,
            "numbered": _add_line_numbers(window_lines, actual_begin)
        }


    def _infer_title_from_context(self, ctx: Dict) -> Dict:
        numbered = "\n".join(ctx.get("numbered", []))
        prompt = f"""
        你是"表格标题/时间/单位抽取助手"。下面是表格**之前**的原文片段（每行前有 Lxxx 行号）。
        任务：找出最可能作为紧随其后表格标题的内容，并同时抽取与该表相关的**最近日期/期间**和**数据单位**。

        判断规则：
        - 表格的标题有可能直接以xx表（其中xx代表任意个字符）的关键字来命名，优先以xx表等作为表格标题（除会企xx表这种除外）
        - 表格的标题也有可能是文章中的各层级标题，标题层级规则：
            - 一级标题：以"一、""二、""三、""四、""五、""六、""七、""八、""九、""十、"开头（中文数字+中文顿号）
            - 二级标题：以"(一)""(二)""(三)""(四)""(五)""(六)""(七)""(八)""(九)""(十)"开头（英文括号包围的中文数字）
            - 三级标题：以"1.""2.""3.""4.""5."开头（阿拉伯数字+英文句点）
            - 四级标题：以"(1)""(2)""(3)""(4)""(5)"开头（英文括号包围的阿拉伯数字）
            - 五级标题：以"1)""2)""3)""4)""5)"开头（右英文括号包围的阿拉伯数字）


        严格要求：
        - 不能编造，所有输出文本必须完全来自片段；
        - 标题若含编号或日期，请去掉日期保留标号；
        - 找不到日期/单位时输出空字符串；
        - 必须记录标题所在的行号（如L3表示第3行）；
        - 当文本中含有会企xx表，会合xx表（xx代表任意长度的字符串或者任意符号），这种不是表格标题，寻找别的文本作为表格标题！
        - 仅输出 JSON。
        - 当片段中出现多级正文标题且无出现xx表等关键字，优先以最靠近片段的正文标题作为表格的标题

        JSON格式：
        {{
          "primary_title": "表格标题（保留标号如'6. 存货'）",
          "title_line": 3,  // 标题所在行号，格式为"xx"
          "level": 0~5,
          "confidence": 0.0~1.0,
          "table_time": "最近日期或期间，如 2022年12月31日",
          "time_line": 5,  // 日期所在行号，找不到则为""
          "unit_name": "如 人民币元 / 人民币万元 / 美元 等，或空串",
          "unit_line": 6   // 单位所在行号，找不到则为""
        }}

        示例：
        如果片段是：
        L1:
        L2:
        L3: 6. 存货
        L4:

        输出应该是：
        {{
          "primary_title": "6. 存货",
          "title_line": 3,
          "level": 3,
          "confidence": 1.0,
          "table_time": "",
          "time_line": "",
          "unit_name": "",
          "unit_line": ""
        }}

        片段（带行号）：
        {numbered}
        """.strip()

        print('【表格标题/时间/单位抽取】')
        print(prompt)
        data = self.llm.call(prompt)
        print('抽取结果：', data)

        # 确保返回的数据包含所有必要字段
        result = {
            "primary_title": data.get("primary_title", ""),
            "title_line": data.get("title_line", ""),
            "level": data.get("level", 0),
            "confidence": data.get("confidence", 0.0),
            "table_time": data.get("table_time", ""),
            "time_line": data.get("time_line", ""),
            "unit_name": data.get("unit_name", ""),
            "unit_line": data.get("unit_line", "")
        }

        # # 如果LLM返回的行号格式是"L3"，转换为实际行号
        # if result["title_line"]:
        #     try:
        #         actual_line_num = int(result["title_line"][1:])
        #         result["title_line_number"] = actual_line_num
        #     except:
        #         result["title_line_number"] = None

        return result


    def _generate_cleaned_table_md(self, df: pd.DataFrame, header_pack: Dict) -> str:
        """根据多级表头解析结果，生成整理后的表格markdown（表头合并为一行）"""
        lines = []

        # 添加合并后的表头（一行）
        resolved_headers = header_pack.get("resolved_headers", [])
        if resolved_headers:
            header_line = "| " + " | ".join(resolved_headers) + " |"
            lines.append(header_line)
            print("777777777788888888888", header_line)  # 调试输出
        else:
            # 如果没有resolved_headers，使用原始第一行
            first_row = df.iloc[0].fillna("").tolist() if not df.empty else []
            header_line = "| " + " | ".join(str(x) for x in first_row) + " |"
            lines.append(header_line)

        # 添加数据行
        data_start = header_pack.get("data_start_line", 1)  # 注意这里改为data_start_line
        start_row0 = max(0, data_start - 1)
        end_row = min(start_row0 + settings.SHEET_PREVIEW_ROWS, len(df))
        data_df = df.iloc[start_row0:end_row, :settings.SHEET_PREVIEW_COLS]

        for _, row in data_df.iterrows():
            cells = [str(x).strip() if x is not None else "" for x in row.tolist()]
            lines.append("| " + " | ".join(cells) + " |")

        md_text = "\n".join(lines)
        print("777777777788888888888", md_text)  # 调试输出
        return md_text


    # ---------- NEW: 并列表识别（基于 Excel 预览 md） ----------
    def _infer_composite_layout_md(self, md_text: str) -> Dict:
        numbered = _add_line_numbers(md_text.splitlines(), 0)
        nl = "\n".join(numbered)

        prompt = f"""
        你是"表格布局分析助手"。以下是一个 Markdown 表格的原文（只看这段，不能编造）：

        {nl}

        任务（严格只从原文判断）：
        1) 判断是否为"复合表格"（多个子表左右并排）。若是，给出子表数量。
        2) 提取每个子表的列头（headers）和列范围（col_range）
           - headers：列头必须逐字来自表头行，保持出现顺序
           - col_range：该子表占用的列范围 [起始列号, 结束列号]（从1开始）

        JSON输出格式：
        {{
          "is_composite": true,
          "subtable_count": 2,
          "subtables": [
            {{
              "id": "T1",
              "headers": ["xx", "xx", "xx", "xx"],
              "col_range": [1, 4]  // 第1-4列属于这个子表
            }},
            {{
              "id": "T2",
              "headers": ["xx", "xx", "xx", "xx"],
              "col_range": [5, 8]  // 第5-8列属于这个子表
            }}
          ],
          "confidence": 0.95
        }}

        重要规则：
        - 若不是复合表：is_composite=false；subtable_count=1；col_range=[1, 总列数]
        - 复合表的判断标准：
          * 存在多个独立的业务主题（如"资产"和"负债"）
          * 列头中间有明显的分界（如重复的列名）
          * 左右两部分可以独立成表
        - col_range必须连续且不重叠，所有子表的列范围合起来应覆盖全部列
        - headers必须逐字来自原文，不可改写

        严格遵守要求：
        -

        示例说明：
        - 如果表格有8列，左边4列是资产相关，右边4列是负债相关
        - 则：T1的col_range=[1,4]，T2的col_range=[5,8]
        """.strip()

        print('22222222222222222', prompt)
        data = self.llm.call(prompt)
        print(data)
        subs = data.get("subtables") or []
        if not subs:
            # 单表（按 md 第一行列数兜底）
            first = md_text.splitlines()[0] if md_text else ""
            col_count = max(1, first.count("|") - 1)
            return {"is_composite": False, "subtables": [{"id": "T1", "name": "", "col_range": [1, col_count]}]}
        # 规范化
        out = []
        for i, st in enumerate(subs, 1):
            rid = st.get("id") or f"T{i}"
            cr = st.get("col_range") or [1, max(1, md_text.splitlines()[0].count("|") - 1)]
            out.append({"id": rid, "name": st.get("name", ""), "col_range": cr})
        return {"is_composite": bool(data.get("is_composite")), "subtables": out}

    def _infer_multilevel_headers_md(self, md_text: str, sub_id: str) -> Dict:
        """修改后的多级表头解析助手"""
        numbered = _add_line_numbers(md_text.splitlines(), 0)
        nl = "\n".join(numbered)

        prompt = f"""
    你是"多级表头解析助手"。你的任务是分析以下Markdown表格，识别哪些行是表头，哪些行是数据内容。

    输入的表格内容（带行号）：
    {nl}

    任务要求：
    1. 识别表头行和数据行
       - 表头行：包含列名称、分类、单位等描述性信息的行
       - 数据行：包含具体数值、具体内容的行

    2. 表头识别规则：
       - 表头可能有多级（1-5行不等）
       - 上级表头可能跨越多列
       - 数据行从第一个包含具体业务数据的行开始
       - 如果该行存在数据，必定不为表头


    3. 多级表头合并规则（重要）：
       - 从上到下逐层连接，用"_"连接
       - 如果上下层内容完全相同，去重只保留一个
       - 保持每一列的完整路径

       举例：
       L1: | 种类 | 期末数 | 期末数 | 期末数 | 期末数 | 期末数 |
       L2: | 种类 | 账面余额 | 账面余额 | 坏账准备 | 坏账准备 |  |
       L3: | 种类 | 金额 | 比例（%） | 金额 | 计提比例（%） |  |

       合并规则说明：
       - 第1列：L1="种类", L2="种类", L3="种类" → 去重后 = "种类"
       - 第2列：L1="期末数", L2="账面余额", L3="金额" → "期末数_账面余额_金额"
       - 第3列：L1="期末数", L2="账面余额", L3="比例（%）" → "期末数_账面余额_比例（%）"
       - 第4列：L1="期末数", L2="坏账准备", L3="金额" → "期末数_坏账准备_金额"
       - 第5列：L1="期末数", L2="坏账准备", L3="计提比例（%）" → "期末数_坏账准备_计提比例（%）"
       - 第6列：L1="期末数", L2="", L3="" → "期末数"（L2和L3中|  |除了字符"|"和空格以外，并无其他字符，如果遇到这种情况，则无需合并L2，L3）
       - 其他列以此类推

    输出JSON格式：
    {{
      "subtable_id": "{sub_id}",
      "header_rows": 3,  // 表头占几行
      "header_lines": [  // 原始表头内容，每行的列值
        ["种类", "期末数", "期末数", "期末数", "期末数", "期末数"],
        ["种类", "账面余额", "账面余额", "坏账准备", "坏账准备", "账面价值"],
        ["种类", "金额", "比例（%）", "金额", "计提比例（%）", "账面价值"]
      ],
      "data_start_line": 4,  // 数据从第几行开始（1-based）
      "resolved_headers": [  // 合并后的列名，必须严格按照上述规则
        "种类",
        "期末数_账面余额_金额",
        "期末数_账面余额_比例（%）",
        "期末数_坏账准备_金额",
        "期末数_坏账准备_计提比例（%）",
        "期末数_账面价值"
      ],
      "column_count": 6,
      "confidence": 0.95
    }}

    关键要求：
    1. 必须保留所有层级的表头信息
    2. 从上到下逐层连接，不能跳过任何层级
    3. 只有当相邻层级内容完全相同时才去重
    4. resolved_headers的数量必须等于表格列数
    5. 如果该行存在||这样的空markdown单元格（||中可能包含多个空格），必定不为表头，举例说明：|xx|  |  |  |xx|  |  |  |，这种行存在|  |（| xx |中的代表任意长度的字符或者符号，|  |代表行中有空数据，两个"|"之间可以存在任意多个空格，或者不存在空格），且该行不存在数值型数据，所以不为表头


    请分析表格并输出JSON：
    """.strip()

        print('【多级表头解析】')
        print(prompt)
        data = self.llm.call(prompt)
        print('解析结果：', data)

        return {
            "subtable_id": data.get("subtable_id", sub_id),
            "header_rows": data.get("header_rows", 1),
            "header_lines": data.get("header_lines", []),
            "data_start_line": data.get("data_start_line", 1),
            "resolved_headers": data.get("resolved_headers", []),
            "column_count": data.get("column_count", 0),
            "confidence": data.get("confidence", 0.0),
        }

    def _textiness_score(self, rows: List[List], col: int) -> float:
        """计算列的“文本密度”：越接近 1 越像名称列"""
        import re
        total = 0
        texty = 0
        for r in rows:
            if col >= len(r):
                continue
            v = r[col]
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            total += 1
            # 数值/百分比/会计格式判定
            num_like = bool(re.fullmatch(r'[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?', s))
            # 含中文/字母视为“文本”
            alpha_like = bool(re.search(r'[\u4e00-\u9fa5A-Za-z]', s))
            if (not num_like) and alpha_like:
                texty += 1
        return (texty / total) if total else 0.0

    def _infer_key_columns_with_llm(
            self, resolved_headers: List[str], data_rows: List[List], sub_id: str,
            sample_per_col: int = 8
    ) -> Dict:
        """
        让 LLM 只做“列角色判定”，输出每列是 item/company/person/value/date/id/other_text/ignore
        然后我们再用程序化方式抽取列值，避免 LLM 漏枚举。
        """

        # 采样：给每列取少量“非空非纯数值”的样本，降低 token 压力
        def norm(x):
            if x is None: return ""
            s = str(x).strip()
            return s

        col_meta = []
        for j, h in enumerate(resolved_headers or []):
            samples = []
            # 修改：对于数值列，也保留一些样本
            for r in data_rows:
                if j >= len(r):
                    continue
                s = norm(r[j])
                if not s:
                    continue
                samples.append(s)  # 修改：保留所有样本（包括数值）
                if len(samples) >= sample_per_col:
                    break
            col_meta.append({"col_index": j, "header": str(h), "samples": samples})

        # prompt = f"""
        # 你是"列角色判定器"。我们给你每一列的列名（已合并后的单行列头）和少量样本值。
        # 请为每列选择 *唯一* 的角色（role）：
        # - item: 财务项目/科目/条目/对象等"关键信息名称列"
        # - company: 公司/单位/客户/供应商等组织名称列
        # - person: 人名列（如法定代表人/负责人/经办人）
        # - value: 金额/数量/比例等数值列（包含纯数字、带千分位、百分比等）
        # - other_text: 以上均不适合，但为可读文本列
        # - ignore: 全为空白或无意义（如注释号、序号等）
        # - time：年份，月份，年月，年末数，期末数等等包含时间的列
        #
        # 判断规则：
        # 1. value角色：样本中70%以上是数值型数据（数字、金额、百分比等）
        # 2. item角色：包含财务科目、项目名称等关键信息
        # 3. company角色：包含常见组织名称结构
        # 4. person角色：包含人名
        # 5. other_text角色：有文本但不属于以上类别
        # 6. ignore角色：空值或无意义列（如全是序号、注释号）
        # 7. time：样本中80%以上是时间类型数据
        #
        # 输出严格 JSON（仅 JSON）：
        # {{
        #   "subtable_id": "{sub_id}",
        #   "columns": [
        #     {{"col_index": 0, "header": "资产", "role": "item", "confidence": 0.95}},
        #     {{"col_index": 1, "header": "注释号", "role": "ignore", "confidence": 0.90}},
        #     {{"col_index": 2, "header": "期末数", "role": "value", "confidence": 0.98}},
        #     {{"col_index": 3, "header": "上年年末数", "role": "value", "confidence": 0.98}}
        #   ],
        #   "key_columns": {{
        #     "item": [0],
        #     "company": [],
        #     "person": [],
        #     "value": [2, 3],
        #     "other_text": [],
        #     "ignore": [1]
        #     "time": [4]
        #   }},
        #   "notes": "第0列为财务项目列，第2、3列为数值列，第1列为注释号"
        # }}
        #
        # 下面是各列的元信息（JSON 数组）：
        # {json.dumps(col_meta, ensure_ascii=False)}
        # """.strip()

        prompt = f"""
        你是"列角色判定器"。我们给你每一列的列名（已合并后的单行列头）和少量样本值。
        请为每列选择 *唯一* 的角色（role）：
        - item: 财务项目/科目/条目/对象等"关键信息名称列"
        - company: 公司/单位/客户/供应商等组织名称列
        - person: 人名列（如法定代表人/负责人/经办人）
        - value: 金额/数量/比例等数值列（包含纯数字、带千分位、百分比等）
        - time: 年份、月份、年月、年末数、期末数、上年数等包含时间的列
        - other_text: 以上均不适合，但为可读文本列
        - ignore: 全为空白或无意义（如注释号、序号等）

        判断规则：
        1. value角色：样本中70%以上是数值型数据（数字、金额、百分比等）
        2. time角色：样本中80%以上是时间类型数据，或列名包含"年"、"月"、"期"、"日期"等时间关键词
        3. item角色：包含财务科目、项目名称等关键信息
        4. company角色：包含组织名称
        5. person角色：包含人名
        6. other_text角色：有文本但不属于以上类别
        7. ignore角色：空值或无意义列（如全是序号、注释号）

        输出严格 JSON（仅 JSON）：
        {{
          "subtable_id": "{sub_id}",
          "key_columns": {{
            "item": [0],
            "company": [],
            "person": [],
            "value": [2, 3],
            "time": [4],
            "other_text": [],
            "ignore": [1]
          }}
        }}

        下面是各列的元信息（JSON 数组）：
        {json.dumps(col_meta, ensure_ascii=False)}
        """.strip()

        data = self.llm.call(prompt) or {}
        # 兜底与清洗
        cols = data.get("columns") or []
        key_cols = data.get("key_columns") or {}
        # 保证所有角色字段存在
        for k in ["item", "company", "person", "value", "time", "other_text", "ignore"]:
            key_cols.setdefault(k, [])

        # 基于“文本密度”做 sanity check：把明显“非文本”的列从 key 中剔除
        for k in ["item", "company", "person", "other_text"]:
            sane_list = []
            for c in key_cols.get(k, []):
                try:
                    score = self._textiness_score(data_rows, int(c))
                    if score >= 0.4:  # 文本密度阈值，过低说明像数值列/空列
                        sane_list.append(int(c))
                except Exception:
                    continue
            key_cols[k] = sane_list

        # 如果所有 key 列都空，用“文本密度最高”的列兜底为 item
        if not any(key_cols.values()):
            scores = [(idx, self._textiness_score(data_rows, idx)) for idx in range(len(resolved_headers or []))]
            scores.sort(key=lambda x: x[1], reverse=True)
            fallback = [i for i, s in scores if s >= 0.6][:1] or ([scores[0][0]] if scores else [])
            key_cols["item"] = fallback

        return {
            "subtable_id": sub_id,
            "key_columns": key_cols
        }


    def _extract_key_information(self, data_rows: List[List], resolved_headers: List[str], sub_id: str) -> Dict:
        """
        新方案：LLM 先判定“哪些列是关键信息列”，然后用代码把这些列里的文本全量收集。
        - 这样不会因为 LLM 漏枚举而漏项
        - 只在“判列”这一步使用 LLM，大幅降低维护成本
        """
        # 1) 让 LLM 判列（仅少量样本），我们再做文本密度兜底
        col_pack = self._infer_key_columns_with_llm(resolved_headers or [], data_rows or [], sub_id=sub_id)
        key_cols = col_pack.get("key_columns", {})

        def norm_cell(x: Optional[str]) -> str:
            if x is None: return ""
            s = str(x).strip()
            # 统一括号/空格等轻量规范化
            trans = str.maketrans({"（": "(", "）": ")", "％": "%", "—": "-", "－": "-"})
            s = s.translate(trans)
            s = re.sub(r'\s+', ' ', s)
            return s

        def is_noise(s: str) -> bool:
            if not s: return True
            # 过滤纯数值/百分比
            if re.fullmatch(r'[-+]?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?', s):
                return True
            # 常见无意义词（尽量保持极小集合，降低维护成本）
            if s in {"合计", "小计", "其中"}:
                return True
            # 过长的一整段描述也过滤（通常不是名称）
            if len(s) > 128:
                return True
            return False

        def collect_from_cols(cols: List[int]) -> List[str]:
            bag = set()
            for r in data_rows or []:
                for c in cols or []:
                    if c >= len(r): continue
                    s = norm_cell(r[c])
                    if is_noise(s):
                        continue
                    bag.add(s)
            # 稳妥：长度升序 + 字典序，便于稳定回放
            out = sorted(bag, key=lambda x: (len(x), x))
            return out

        # 2) 程序化全量收集
        items = collect_from_cols(key_cols.get("item", []))
        company = collect_from_cols(key_cols.get("company", []))
        person = collect_from_cols(key_cols.get("person", []))
        other = collect_from_cols(key_cols.get("other_text", []))

        # ✅ 简化：只打印key_columns，不生成summary
        print(f'【列角色判定】{sub_id}: {key_cols}')

        return {
            "subtable_id": sub_id,
            "key_information": {
                "财务指标": items,
                "公司名称": company,
                "人名": person,
                "其他关键项": other
            },
            "key_columns": key_cols
        }


    # def extract_from_xlsx_mentions(self, text: str) -> List[Dict]:

    def extract_from_xlsx_mentions(self, text: str,
                                       input_file_path: str = "",
                                       agent_user_id: str = "",
                                       task_id: str = "") -> List[Dict]:

        """
        修正后的主函数，支持路径参数

        Args:
            text: 文档文本
            input_file_path: 基础输入路径
            agent_user_id: 用户ID
            task_id: 任务ID
        """
        lines = text.splitlines()
        hits = find_xlsx_mentions(text)
        tables: List[Dict] = []

        for hit in hits:
            file_name = hit["file_name"]
            line_idx = hit["line_idx"]

            file_base_name = os.path.splitext(file_name)[0]

            # 1) 标题/时间/单位
            ctx = self._get_context_before(lines, line_idx)
            title_pack = self._infer_title_from_context(ctx)
            title = title_pack.get("primary_title") or "未命名表格"

            # 2) 附件定位/落盘 - 传入路径参数
            staged = resolve_and_stage_attachment(
                file_name,
                input_base_path=input_file_path,
                agent_user_id=agent_user_id,
                task_id=task_id
            )

            if not staged:
                # 文件不存在的处理逻辑
                table_id = generate_table_id(file_name, title)
                tables.append({
                    "table_id": table_id,
                    "start_line": line_idx + 1,
                    "end_line": line_idx + 1,
                    "title": title,
                    "title_line": title_pack.get("title_line", ""),
                    "title_confidence": float(title_pack.get("confidence", 0.0)),
                    "table_time": title_pack.get("table_time", ""),
                    "time_line": title_pack.get("time_line", ""),
                    "unit_name": title_pack.get("unit_name", ""),
                    "unit_line": title_pack.get("unit_line", ""),
                    "file": file_name,
                    "file_path": None,
                    "subtables": [],
                    "note": f"未找到该文件"
                })
                continue

            # 生成表ID（使用file_path + title）
            table_id = generate_table_id(staged, title)  # 新增

            # 3) 逐工作表解析
            wb = load_workbook(staged, data_only=True)
            subtables_out = []

            for si, sname in enumerate(wb.sheetnames):
                df = read_xlsx_expand_merged(staged, sname)
                print('我是读取的excel内容：：：：：：：：：：：：：', df)
                # 生成sheet标识：文件名_Sheet1 或 文件名_Sheet名称
                sheet_id = f"{file_base_name}_Sheet{si + 1}"  # 例如: "excel_table_7_Sheet1"

                # 3.1 先生成原始md预览
                md_lines = df_to_md_lines(df, settings.SHEET_PREVIEW_ROWS, settings.SHEET_PREVIEW_COLS)
                md_text = "\n".join(md_lines)

                # 3.2 【新顺序】先进行多级表头解析
                # 传入包含文件名的sheet_id
                header_pack = self._infer_multilevel_headers_md(md_text, sub_id=sheet_id)

                # 3.3 生成整理后的表格（表头合并为一行）
                cleaned_md_text = self._generate_cleaned_table_md(df, header_pack)

                # 3.4 【新顺序】再进行布局分析（基于整理后的表格）
                layout = self._infer_composite_layout_md(cleaned_md_text)
                subs = layout["subtables"] if layout.get("subtables") else [{"id": "T1", "col_range": [1, df.shape[1]]}]

                # 3.5 处理每个子表
                for st in subs:
                    # 生成完整的子表ID：文件名_Sheet编号:子表编号
                    if layout.get("is_composite", False) and len(subs) > 1:
                        sub_id = f"{file_base_name}_Sheet{si + 1}_{st.get('id', 'T1')}"
                        # 例如: "excel_table_7_Sheet1:T1" （如果是复合表）
                    else:
                        sub_id = f"{file_base_name}_Sheet{si + 1}"
                        # 例如: "excel_table_7_Sheet1" （如果是单表）

                    col_range = st.get("col_range", [1, df.shape[1]])

                    # 获取子表数据
                    sub_df = slice_df_columns(df, col_range)
                    data_start = header_pack.get("data_start_line", 1)

                    # 从数据开始行提取数据
                    start_row0 = max(0, data_start - 1)
                    data_df = sub_df.iloc[start_row0:, :].copy()

                    def norm(x):
                        if x is None: return None
                        s = str(x).strip()
                        return None if s == "" or s.lower() == "null" else s

                    data_rows = [[norm(x) for x in row] for row in data_df.values.tolist()]

                    # 获取该子表对应的列头
                    resolved_headers = header_pack.get("resolved_headers", [])
                    sub_headers = resolved_headers[col_range[0] - 1:col_range[1]] if resolved_headers else []

                    # 【新增】提取关键信息
                    key_info = self._extract_key_information(data_rows, sub_headers, sub_id)

                    subtables_out.append({
                        "sheet_name": sname,  # 保留原始sheet名称
                        "sheet_index": si + 1,  # 新增：sheet序号
                        "subtable_id": sub_id,  # 使用包含文件名的ID
                        "name": st.get("name", ""),
                        "col_range": col_range,
                        "resolved_headers": sub_headers,
                        "header_rows": header_pack.get("header_rows", 1),
                        "header_lines": header_pack.get("header_lines", []),
                        "header_confidence": header_pack.get("confidence", 0.0),
                        "data_start_line": data_start,
                        "data_rows": data_rows,
                        "is_composite": layout.get("is_composite", False),
                        # 【新增】关键信息字段
                        "key_information": key_info.get("key_information", {}),
                        "key_columns": key_info.get("key_columns", {})
                    })

            tables.append({
                "table_id": table_id,  # 新增：添加在第一个位置
                "start_line": line_idx + 1,
                "end_line": line_idx + 1,
                "title": title,
                "title_line": title_pack.get("title_line", ""),
                # "title_line_number": title_pack.get("title_line_number"),
                "title_confidence": float(title_pack.get("confidence", 0.0)),
                "table_time": title_pack.get("table_time", ""),
                "time_line": title_pack.get("time_line", ""),
                "unit_name": title_pack.get("unit_name", ""),
                "unit_line": title_pack.get("unit_line", ""),
                "file": file_name,
                "file_base_name": file_base_name,  # 新增：文件基础名称
                "file_path": staged,
                "subtables": subtables_out
            })

        return tables


    # ---------- 输出 ----------
    def save_results(self, tables: List[Dict], filename: Optional[str] = None):
        """
        保存表格提取结果

        Args:
            tables: 提取的表格列表
            filename: 输出文件名（可选）

        Returns:
            str: 保存文件的完整绝对路径
        """
        if filename is None:
            filename = f"tables_{_now_suffix()}"
        else:
            # 如果提供了filename，添加 _tables 后缀
            base_name = os.path.splitext(filename)[0]  # 去除可能的扩展名
            filename = f"{base_name}_tables"

        # 确保文件名以.json结尾
        if not filename.endswith('.json'):
            filename = f"{filename}.json"

        path = os.path.join(self.s.OUTPUT_DIR, filename)
        # 转换为绝对路径
        path = os.path.abspath(path)

        # 直接覆盖旧文件（不需要检查是否存在）
        with open(path, "w", encoding="utf-8") as f:
            json.dump({"tables": tables, "count": len(tables)}, f, ensure_ascii=False, indent=2)

        print(f"[OK] 结果已保存：{path}")
        return path  # 返回完整的绝对路径


    def format_preview(self, tables: List[Dict]) -> str:
        out = []
        for ti, t in enumerate(tables, 1):
            out.append("=" * 70)
            out.append(f"表 {ti}  行号: L{t['start_line']} - L{t['end_line']}")
            out.append(f"标题: {t.get('title', '')}")
            if t.get("table_time"): out.append(f"时间: {t['table_time']}")
            if t.get("unit_name"):  out.append(f"单位: {t['unit_name']}")
            out.append(f"文件: {t.get('file')}  → {t.get('file_path')}")

            for st in t.get("subtables", []):
                out.append(f"  ├─ [{st.get('sheet_name')}] 子表 {st.get('subtable_id')}: 列范围 {st.get('col_range')}")
                headers = st.get("resolved_headers", [])
                if headers:
                    out.append("  │  列名: " + " | ".join(headers))
                out.append(f"  │  数据行数: {len(st.get('data_rows', []))}")

                # ✅ 删除 key_info_summary 的显示
                # 改为直接显示 key_columns
                key_cols = st.get("key_columns", {})
                if key_cols:
                    out.append(f"  │  列角色: {key_cols}")

                # 保留 key_information 的显示
                key_info = st.get("key_information", {})
                if key_info:
                    for category, items in key_info.items():
                        if items:
                            out.append(f"  │    - {category}: {', '.join(items[:5])}{'...' if len(items) > 5 else ''}")

        return "\n".join(out)


# =========================
# 示例 main（改为从 xlsx 引用解析）
# =========================
def main(text, json_file_name=None, output_dir=None,
         input_file_path=None, agent_user_id=None, task_id=None):
    """
    主处理函数

    Args:
        text: 文档文本内容
        json_file_name: 输出的JSON文件名（不含扩展名）
        output_dir: 自定义输出目录（可选）
        input_file_path: 输入文件基础路径（新增）
        agent_user_id: 用户ID（新增）
        task_id: 任务ID（新增）

    Returns:
        str: 生成的tables.json文件的完整绝对路径
    """
    # 更新全局配置
    if input_file_path:
        settings.INPUT_FILE_PATH = input_file_path
    if agent_user_id:
        settings.AGENT_USER_ID = str(agent_user_id)
    if task_id:
        settings.TASK_ID = str(task_id)

    # 如果指定了输出目录，临时覆盖配置
    original_output_dir = None
    original_attachment_dir = None

    if output_dir:
        original_output_dir = settings.OUTPUT_DIR
        original_attachment_dir = settings.ATTACHMENT_OUT_DIR

        settings.OUTPUT_DIR = output_dir
        settings.ATTACHMENT_OUT_DIR = output_dir

        os.makedirs(settings.OUTPUT_DIR, exist_ok=True)
        os.makedirs(settings.ATTACHMENT_OUT_DIR, exist_ok=True)

        print(f"✅ 使用自定义输出目录: {output_dir}")

    # 打印配置信息
    if input_file_path and agent_user_id and task_id:
        full_input_path = os.path.join(input_file_path, str(agent_user_id), str(task_id))
        print("=" * 80)
        print("配置信息:")
        print(f"  输入路径基础: {input_file_path}")
        print(f"  用户ID: {agent_user_id}")
        print(f"  任务ID: {task_id}")
        print(f"  完整输入路径: {full_input_path}")
        print(f"  输出目录: {settings.OUTPUT_DIR}")
        print("=" * 80)

    try:
        extractor = MultiHeaderTableExtractor()

        # 传入路径参数
        tables = extractor.extract_from_xlsx_mentions(
            text,
            input_file_path=input_file_path or "",
            agent_user_id=agent_user_id or "",
            task_id=task_id or ""
        )

        print(extractor.format_preview(tables))
        output_path = extractor.save_results(tables, filename=json_file_name)
        return output_path

    finally:
        # 恢复原始输出目录设置
        if original_output_dir is not None:
            settings.OUTPUT_DIR = original_output_dir
        if original_attachment_dir is not None:
            settings.ATTACHMENT_OUT_DIR = original_attachment_dir


if __name__ == "__main__":
    # 把你“全文 + xlsx 名称”的文本放进来（示例已含 excel_table_1.xlsx 等占位）
    text = r"""
1.1．资质证书表及证明材料
table/XA_certificate_0_table_1.xlsx
1.1.1.IS09001质量管理体系认证证书（有效期到2027年7
月2日)
image/XA_certificate_1_layout_det_res_1.png
        """
    # ========== 配置参数 ==========
    # 输入文件基础路径
    input_file_path = "/home/cqj/my-doc-system-uploads/save/"

    # 用户ID和任务ID
    agent_user_id = 1001
    task_id = 2

    # 输出文件名
    json_file_name = "证书文档"

    # 自定义输出目录
    custom_output_dir = "/data/cwd_cq/out"
    # ==============================

    # 调用main函数，传入所有参数
    tables_path = main(
        text,
        json_file_name=json_file_name,
        output_dir=custom_output_dir,
        input_file_path=input_file_path,
        agent_user_id=agent_user_id,
        task_id=task_id
    )

    print(f"\n✅ 处理完成！")
    print(f"输出文件: {tables_path}")

